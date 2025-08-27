from flask import Flask, render_template, request, redirect, url_for, jsonify, session, flash, send_file
import pandas as pd
from datetime import datetime, date
import os
import hashlib
from werkzeug.utils import secure_filename
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = 'associacao_amigo_do_povo_2024_secure_key'

# Usuários do sistema
USUARIOS = {
    'admin': {
        'senha': hashlib.sha256('admin123'.encode()).hexdigest(),
        'nome': 'Administrador',
        'nivel': 'admin'
    },
    'usuario': {
        'senha': hashlib.sha256('usuario123'.encode()).hexdigest(), 
        'nome': 'Usuário',
        'nivel': 'usuario'
    }
}

# Configuração para upload
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Atividades disponíveis
ATIVIDADES = ['Natação', 'Informática', 'Fisioterapia', 'Dança', 'Hidroginastica', 
              'Funcional', 'Karatê', 'Bombeiro mirim', 'Capoeira']

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def verificar_login():
    return 'usuario_logado' in session

def login_obrigatorio(f):
    def wrapper(*args, **kwargs):
        if not verificar_login():
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

def gerar_planilha_frequencia(atividade, mes_ano='2024-12'):
    """Gera planilha de frequência modelo"""
    try:
        ano, mes = mes_ano.split('-')
        mes_nome = {
            '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
            '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
            '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
        }.get(mes, 'Dezembro')
        
        # Dados de exemplo baseados na atividade
        alunos_exemplo = {
            'Natação': [
                {'nome': 'Ana Silva', 'telefone': '(11) 99999-1111', 'turma': 'Manhã'},
                {'nome': 'João Santos', 'telefone': '(11) 99999-2222', 'turma': 'Tarde'},
                {'nome': 'Maria Oliveira', 'telefone': '(11) 99999-3333', 'turma': 'Manhã'},
            ],
            'Informática': [
                {'nome': 'Pedro Costa', 'telefone': '(11) 99999-4444', 'turma': 'Básico'},
                {'nome': 'Carla Lima', 'telefone': '(11) 99999-5555', 'turma': 'Avançado'},
                {'nome': 'Rafael Souza', 'telefone': '(11) 99999-6666', 'turma': 'Básico'},
            ]
        }
        
        alunos = alunos_exemplo.get(atividade, [
            {'nome': f'Aluno {i+1}', 'telefone': f'(11) 9999-{i+1:04d}', 'turma': 'Turma A'} 
            for i in range(5)
        ])
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = mes_nome
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Cabeçalho principal
        ws.merge_cells('A1:AH1')
        ws['A1'] = f'CONTROLE DE FREQUÊNCIA - {atividade.upper()} - {mes_nome.upper()} {ano}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Cabeçalhos das colunas
        ws['A3'] = 'ALUNO'
        ws['B3'] = 'TELEFONE'
        ws['C3'] = 'TURMA'
        
        # Aplicar estilo aos cabeçalhos
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = center_alignment
        
        # Dias do mês (D3 até AH3 = 31 dias)
        for dia in range(1, 32):
            if dia <= 26:
                col_letter = chr(67 + dia)  # D=68, E=69, etc.
            else:
                col_letter = 'A' + chr(67 + dia - 26)
                
            ws[f'{col_letter}3'] = dia
            ws[f'{col_letter}3'].font = header_font
            ws[f'{col_letter}3'].fill = header_fill
            ws[f'{col_letter}3'].alignment = center_alignment
        
        # Adicionar alunos
        for idx, aluno in enumerate(alunos, start=4):
            ws[f'A{idx}'] = aluno['nome']
            ws[f'B{idx}'] = aluno['telefone']
            ws[f'C{idx}'] = aluno['turma']
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Ajustar largura dos dias
        for dia in range(1, 32):
            if dia <= 26:
                col_letter = chr(67 + dia)
            else:
                col_letter = 'A' + chr(67 + dia - 26)
            ws.column_dimensions[col_letter].width = 3
        
        # Adicionar instruções
        linha_instrucao = len(alunos) + 6
        ws[f'A{linha_instrucao}'] = 'INSTRUÇÕES:'
        ws[f'A{linha_instrucao}'].font = Font(bold=True)
        ws[f'A{linha_instrucao + 1}'] = 'P = Presente | F = Falta | J = Justificado | X = Presente'
        ws[f'A{linha_instrucao + 2}'] = 'Deixe em branco se não houve aula no dia'
        
        return wb
        
    except Exception as e:
        print(f"Erro ao gerar planilha: {e}")
        return None

# ROTAS
@app.route('/')
def index():
    if not verificar_login():
        return render_template('splash.html')
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        senha = request.form.get('senha')
        
        if usuario in USUARIOS:
            senha_hash = hashlib.sha256(senha.encode()).hexdigest()
            if USUARIOS[usuario]['senha'] == senha_hash:
                session['usuario_logado'] = usuario
                session['usuario_nome'] = USUARIOS[usuario]['nome']
                session['usuario_nivel'] = USUARIOS[usuario]['nivel']
                flash('Login realizado com sucesso!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Senha incorreta!', 'error')
        else:
            flash('Usuário não encontrado!', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logout realizado com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_obrigatorio
def dashboard():
    usuario_nome = session.get('usuario_nome', 'Usuário')
    stats = {
        'total_alunos': 439,
        'presencas_hoje': 25,
        'total_registros': 1250,
        'alunos_ativos': 420
    }
    return render_template('dashboard.html', stats=stats, presencas_hoje=[], usuario_nome=usuario_nome)

@app.route('/backup_planilhas')
@login_obrigatorio
def backup_planilhas():
    if session.get('usuario_nivel') != 'admin':
        flash('Acesso negado! Apenas administradores podem acessar esta área.', 'error')
        return redirect(url_for('dashboard'))
    
    usuario_nome = session.get('usuario_nome', 'Usuário')
    return render_template('backup_planilhas.html', 
                         atividades=ATIVIDADES, 
                         usuario_nome=usuario_nome)

@app.route('/gerar_planilha/<atividade>')
@login_obrigatorio
def gerar_planilha_route(atividade):
    if session.get('usuario_nivel') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        mes_ano = request.args.get('mes_ano', '2024-12')
        wb = gerar_planilha_frequencia(atividade, mes_ano)
        
        if not wb:
            return jsonify({'error': f'Erro ao gerar planilha para {atividade}'}), 400
        
        # Criar buffer em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo
        ano, mes = mes_ano.split('-')
        mes_nome = {
            '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
            '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
            '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
        }.get(mes, 'Dezembro')
        
        filename = f'Frequencia_{atividade.replace(" ", "_")}_{mes_nome}_{ano}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload_planilha', methods=['POST'])
@login_obrigatorio
def upload_planilha():
    if session.get('usuario_nivel') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        if 'arquivo' not in request.files:
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        file = request.files['arquivo']
        atividade = request.form.get('atividade')
        
        if file.filename == '' or not atividade:
            return jsonify({'error': 'Arquivo e atividade são obrigatórios'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{timestamp}_{atividade}_{filename}'
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            return jsonify({'success': True, 'message': f'Planilha de {atividade} enviada com sucesso!'})
        
        return jsonify({'error': 'Tipo de arquivo não permitido'}), 400
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/listar_backups')
@login_obrigatorio
def listar_backups():
    if session.get('usuario_nivel') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        backups = []
        
        if os.path.exists(UPLOAD_FOLDER):
            for filename in os.listdir(UPLOAD_FOLDER):
                if filename.endswith(('.xlsx', '.xls', '.csv')):
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    stat = os.stat(filepath)
                    
                    backups.append({
                        'nome': filename,
                        'tamanho': f'{stat.st_size / 1024:.1f} KB',
                        'modificado': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
                        'atividade': filename.split('_')[1] if '_' in filename else 'Desconhecida'
                    })
        
        backups.sort(key=lambda x: x['modificado'], reverse=True)
        return jsonify(backups)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
